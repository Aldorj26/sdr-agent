#!/usr/bin/env python3
"""
Gera planilha XLSX com 3 abas (ações Campanha AIVA 2026-05-11):
- A_Cadastro_Completo: 13 leads pra Nei mover pra "Em Análise CAF"
- B_Aguardando_Aprovacao: 25 leads pra Eduardo aprovar
- C_Atendimento_Humano: 117 leads marcados acionar_humano=true (priorizado por fase)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── DADOS ────────────────────────────────────────────────────────────────────

LISTA_A = [
    ("Eletrocell", "5551982435754", "2068", "17/04/2026 00:56", 24),
    ("Club infocell 1 e 2 / Adonai cell", "5511933011098", "2557", "27/04/2026 15:11", 13),
    ("Smarting", "5554999971709", "2388", "27/04/2026 16:33", 13),
    ("BibiCelulares", "5516994358090", "3340", "28/04/2026 14:24", 12),
    ("Lojas Moreira", "5543996076442", "1859", "28/04/2026 20:43", 12),
    ("AppleCel Celulares", "5551999584395", "2334", "29/04/2026 11:59", 12),
    ("Calado Celulares e MM Celulares", "5517996326605", "3386", "29/04/2026 14:42", 11),
    ("Lfn Celulares", "5512983118100", "2575", "29/04/2026 14:42", 11),
    ("Matheus Roney Barbosa da Rosa", "5551991780868", "2107", "30/04/2026 17:12", 10),
    ("Deivid Celulares", "5527999954352", "3678", "04/05/2026 11:40", 7),
    ("Renan Celulares", "5519987613584", "3483", "05/05/2026 10:51", 6),
    ("A.S Celular", "5532999957492", "4683", "06/05/2026 10:57", 5),
    ("Eletrocel", "5533999496090", "4722", "06/05/2026 14:26", 4),
]

LISTA_B = [
    ("Loja", "5519991411740", "1764", "22/04/2026 17:12", 18),
    ("DaTell Celulares", "5554996333396", "2558", "27/04/2026 11:44", 14),
    ("Sucatao96", "5521996969697", "2040", "27/04/2026 12:10", 14),
    ("EB IMPORTS / EB acessórios e celular", "5515998501359", "2690", "27/04/2026 14:32", 13),
    ("Jv Cell", "5514996353235", "2644", "27/04/2026 16:53", 13),
    ("Maia Celulares", "5517991794911", "3374", "28/04/2026 10:41", 13),
    ("JPINFOCEL", "5522981361677", "3496", "29/04/2026 10:05", 12),
    ("Tim acessórios", "5519978252899", "3449", "04/05/2026 10:08", 7),
    ("Bew Celulares", "5527999013469", "3645", "04/05/2026 12:44", 7),
    ("Mercado do Celular", "5531975124100", "4521", "05/05/2026 11:06", 6),
    ("Tech Up", "5531993100036", "4598", "05/05/2026 14:56", 5),
    ("Pixel Games", "5543991232205", "1850", "06/05/2026 17:00", 4),
    ("Parcial Celular", "5537998051419", "4867", "07/05/2026 14:57", 3),
    ("Cristal celular", "5534999020003", "4797", "07/05/2026 16:37", 3),
    ("TOP Celulares", "5535999760808", "4851", "07/05/2026 18:41", 3),
    ("GR Celular", "5538999028082", "4938", "08/05/2026 09:35", 3),
    ("smarttech", "5554997054975", "4957", "11/05/2026 11:17", 0),
    ("CASE CELULARES", "5531985580850", "4544", "11/05/2026 11:39", 0),
    ("Fix Info", "5542999496353", "2704", "11/05/2026 12:27", 0),
    ("King Cell", "5519998895943", "2290", "11/05/2026 12:29", 0),
    ("Coimbra Celular", "5533988970329", "4695", "11/05/2026 12:32", 0),
    ("Liana Celulares", "5528999852730", "3701", "11/05/2026 12:43", 0),
    ("Nane Cell assistência técnica e tecnologia", "5547997053334", "2705", "11/05/2026 13:05", 0),
    ("Ricardo Celulares", "5555996833493", "4986", "11/05/2026 13:11", 0),
    ("JF CELULARES", "5541999053122", "5080", "11/05/2026 13:46", 0),
]

LISTA_C = [
    # CADASTRO_COMPLETO (13)
    ("Eletrocell", "5551982435754", "CADASTRO_COMPLETO", "2068", "17/04/2026 00:56", 24),
    ("Club infocell 1 e 2 / Adonai cell", "5511933011098", "CADASTRO_COMPLETO", "2557", "27/04/2026 15:11", 13),
    ("Smarting", "5554999971709", "CADASTRO_COMPLETO", "2388", "27/04/2026 16:33", 13),
    ("BibiCelulares", "5516994358090", "CADASTRO_COMPLETO", "3340", "28/04/2026 14:24", 12),
    ("Lojas Moreira", "5543996076442", "CADASTRO_COMPLETO", "1859", "28/04/2026 20:43", 12),
    ("AppleCel Celulares", "5551999584395", "CADASTRO_COMPLETO", "2334", "29/04/2026 11:59", 12),
    ("Calado Celulares e MM Celulares", "5517996326605", "CADASTRO_COMPLETO", "3386", "29/04/2026 14:42", 11),
    ("Lfn Celulares", "5512983118100", "CADASTRO_COMPLETO", "2575", "29/04/2026 14:42", 11),
    ("Matheus Roney Barbosa da Rosa", "5551991780868", "CADASTRO_COMPLETO", "2107", "30/04/2026 17:12", 10),
    ("Deivid Celulares", "5527999954352", "CADASTRO_COMPLETO", "3678", "04/05/2026 11:40", 7),
    ("Renan Celulares", "5519987613584", "CADASTRO_COMPLETO", "3483", "05/05/2026 10:51", 6),
    ("A.S Celular", "5532999957492", "CADASTRO_COMPLETO", "4683", "06/05/2026 10:57", 5),
    ("Eletrocel", "5533999496090", "CADASTRO_COMPLETO", "4722", "06/05/2026 14:26", 4),
    # COLETANDO_COMPLEMENTO (5)
    ("Rede HG Smart", "5551991441615", "COLETANDO_COMPLEMENTO", "2104", "27/04/2026 15:44", 13),
    ("Trackcel", "5547991402012", "COLETANDO_COMPLEMENTO", "3567", "01/05/2026 17:22", 9),
    ("Balcão de testes, Planeta celular, Digital eletrônicos", "5531991439151", "COLETANDO_COMPLEMENTO", "4580", "05/05/2026 11:25", 6),
    ("Esmeraldas Celulares", "5531994076919", "COLETANDO_COMPLEMENTO", "4606", "05/05/2026 11:26", 6),
    ("SOS celulares", "5538988030034", "COLETANDO_COMPLEMENTO", "4894", "07/05/2026 15:54", 3),
    # AGUARDANDO_APROVACAO (8)
    ("Loja", "5519991411740", "AGUARDANDO_APROVACAO", "1764", "22/04/2026 17:12", 18),
    ("EB IMPORTS / EB acessórios e celular", "5515998501359", "AGUARDANDO_APROVACAO", "2690", "27/04/2026 14:32", 13),
    ("Jv Cell", "5514996353235", "AGUARDANDO_APROVACAO", "2644", "27/04/2026 16:53", 13),
    ("Tim acessórios", "5519978252899", "AGUARDANDO_APROVACAO", "3449", "04/05/2026 10:08", 7),
    ("Bew Celulares", "5527999013469", "AGUARDANDO_APROVACAO", "3645", "04/05/2026 12:44", 7),
    ("Tech Up", "5531993100036", "AGUARDANDO_APROVACAO", "4598", "05/05/2026 14:56", 5),
    ("Cristal celular", "5534999020003", "AGUARDANDO_APROVACAO", "4797", "07/05/2026 16:37", 3),
    ("CASE CELULARES", "5531985580850", "AGUARDANDO_APROVACAO", "4544", "11/05/2026 11:39", 0),
    # INTERESSADO (13)
    ("Loja", "5541998574000", "INTERESSADO", "1785", "28/04/2026 14:31", 12),
    ("Loja", "5545999354090", "INTERESSADO", "2003", "29/04/2026 11:36", 12),
    ("Loja", "555192146639", "INTERESSADO", "3565", "04/05/2026 10:00", 7),
    ("Loja", "554784196636", "INTERESSADO", "3563", "04/05/2026 10:00", 7),
    ("Loja", "554163474679", "INTERESSADO", "", "04/05/2026 10:00", 7),
    ("Loja", "554796613449", "INTERESSADO", "3562", "04/05/2026 10:00", 7),
    ("Glocall", "5519981390752", "INTERESSADO", "3453", "04/05/2026 10:00", 7),
    ("Loja", "5519983135691", "INTERESSADO", "3473", "04/05/2026 10:00", 7),
    ("Loja", "5511962793765", "INTERESSADO", "", "06/05/2026 18:05", 4),
    ("Loja", "5514997289029", "INTERESSADO", "2650", "07/05/2026 10:00", 4),
    ("Loja", "5514996142915", "INTERESSADO", "4650", "11/05/2026 10:07", 0),
    ("Mataruco iPhones", "5544999180509", "INTERESSADO", "1947", "11/05/2026 12:23", 0),
    ("Loja", "5527997991171", "INTERESSADO", "3622", "11/05/2026 12:29", 0),
    # AGUARDANDO (24)
    ("DESTIN CEL", "554797862194", "AGUARDANDO", "1804", "10/04/2026 11:12", 31),
    ("Loja", "5542999561960", "AGUARDANDO", "1824", "10/04/2026 11:38", 31),
    ("Loja", "5545999414795", "AGUARDANDO", "2005", "10/04/2026 11:47", 31),
    ("JS celulares Ltda", "5545998260990", "AGUARDANDO", "1992", "11/04/2026 12:13", 30),
    ("Hit One", "5551985848584", "AGUARDANDO", "2091", "14/04/2026 11:04", 27),
    ("Loja", "5551996032091", "AGUARDANDO", "2135", "27/04/2026 15:38", 13),
    ("Loja", "5519981863879", "AGUARDANDO", "3463", "29/04/2026 09:24", 12),
    ("Loja", "5542988123016", "AGUARDANDO", "1801", "04/05/2026 10:12", 7),
    ("Clinicell", "5544999646486", "AGUARDANDO", "1959", "04/05/2026 10:14", 7),
    ("Loja", "5527997300749", "AGUARDANDO", "3611", "04/05/2026 11:11", 7),
    ("Loja", "5519988500914", "AGUARDANDO", "3487", "05/05/2026 10:17", 6),
    ("Loja", "5531980401412", "AGUARDANDO", "4525", "05/05/2026 11:05", 6),
    ("Loja", "5532984885157", "AGUARDANDO", "4662", "06/05/2026 09:50", 5),
    ("Smarting", "5554991355935", "AGUARDANDO", "2370", "06/05/2026 10:06", 5),
    ("Loja", "5551981326846", "AGUARDANDO", "2058", "11/05/2026 12:23", 0),
    ("Loja", "5511948515023", "AGUARDANDO", "2218", "11/05/2026 12:24", 0),
    ("Loja", "5511947905670", "AGUARDANDO", "2217", "11/05/2026 12:24", 0),
    ("Loja", "5511940334116", "AGUARDANDO", "2201", "11/05/2026 12:24", 0),
    ("Loja", "5512996776392", "AGUARDANDO", "2584", "11/05/2026 12:26", 0),
    ("Loja", "5513991787703", "AGUARDANDO", "2614", "11/05/2026 12:26", 0),
    ("Loja", "5527999121898", "AGUARDANDO", "4513", "11/05/2026 12:30", 0),
    ("Zoom Celulares", "5554996291945", "AGUARDANDO", "4954", "11/05/2026 12:33", 0),
    ("Loja", "5543999158459", "AGUARDANDO", "1880", "11/05/2026 12:36", 0),
    ("Loja", "5534998402851", "AGUARDANDO", "4783", "11/05/2026 12:46", 0),
    # FORMULARIO_ENVIADO legacy (21)
    ("Daniel Tomio tech", "554797738268", "FORMULARIO_ENVIADO", "", "09/04/2026 15:53", 31),
    ("Cell Consert Celulares", "5541999180808", "FORMULARIO_ENVIADO", "1796", "09/04/2026 16:32", 31),
    ("Loja", "5541999186723", "FORMULARIO_ENVIADO", "1776", "09/04/2026 20:12", 31),
    ("Loja", "5511975759781", "FORMULARIO_ENVIADO", "1791", "09/04/2026 21:00", 31),
    ("NANO CELL LTDA", "5543984881652", "FORMULARIO_ENVIADO", "1844", "10/04/2026 11:51", 31),
    ("Ponto do Celular", "5545998022122", "FORMULARIO_ENVIADO", "1989", "10/04/2026 12:11", 31),
    ("R Castoldi Celulares", "5544998572367", "FORMULARIO_ENVIADO", "1932", "10/04/2026 12:18", 31),
    ("Talau Celulares", "5546999402773", "FORMULARIO_ENVIADO", "2037", "11/04/2026 11:35", 30),
    ("teste idempotencia ignorar", "5543988208134", "FORMULARIO_ENVIADO", "2039", "11/04/2026 14:08", 29),
    ("Unique", "5546991012121", "FORMULARIO_ENVIADO", "2020", "13/04/2026 11:12", 28),
    ("Kw celulares", "5543998051903", "FORMULARIO_ENVIADO", "2043", "13/04/2026 14:50", 27),
    ("Up Cell", "5551994212111", "FORMULARIO_ENVIADO", "2124", "14/04/2026 11:07", 27),
    ("Sandex Celulares", "5551984251423", "FORMULARIO_ENVIADO", "2077", "14/04/2026 11:16", 27),
    ("Nei Teste", "5548999155655", "FORMULARIO_ENVIADO", "2190", "14/04/2026 23:19", 26),
    ("FixxCell", "554588225292", "FORMULARIO_ENVIADO", "2289", "15/04/2026 09:41", 26),
    ("TIM", "5511981180100", "FORMULARIO_ENVIADO", "2270", "15/04/2026 12:07", 26),
    ("Smart Celulares - Lapa", "5511978543995", "FORMULARIO_ENVIADO", "2265", "15/04/2026 12:07", 26),
    ("Loja", "5551996180000", "FORMULARIO_ENVIADO", "2139", "15/04/2026 13:30", 26),
    ("Elder Tech", "5545988139336", "FORMULARIO_ENVIADO", "1974", "15/04/2026 17:23", 25),
    ("Smart Power", "5551991188564", "FORMULARIO_ENVIADO", "2100", "15/04/2026 21:47", 25),
    ("PhoneTech Celulares", "5554996023856", "FORMULARIO_ENVIADO", "2384", "16/04/2026 14:49", 24),
    ("FM Celulares", "5551998625481", "FORMULARIO_ENVIADO", "2324", "16/04/2026 15:28", 24),
    # DESCARTADO/BOT_DETECTADO + outros (33)
    ("Loja", "5519997500329", "DESCARTADO", "1763", "23/04/2026 09:03", 18),
    ("Loja", "5519989430237", "DESCARTADO", "1760", "23/04/2026 09:04", 18),
    ("Lojista", "5511987231799", "BOT_DETECTADO", "2446", "24/04/2026 09:48", 17),
    ("Lojista", "5511990253049", "BOT_DETECTADO", "2470", "24/04/2026 09:48", 17),
    ("Lojista", "5511995859298", "BOT_DETECTADO", "2510", "24/04/2026 09:50", 17),
    ("Sos Celulares", "5554991675561", "BOT_DETECTADO", "2376", "25/04/2026 14:28", 15),
    ("Prime Celulares e Acessórios", "5543996625130", "BOT_DETECTADO", "1866", "25/04/2026 15:02", 15),
    ("Loja", "5512982679030", "BOT_DETECTADO", "2571", "27/04/2026 11:26", 14),
    ("Loja", "5514982178941", "BOT_DETECTADO", "2633", "27/04/2026 11:35", 14),
    ("Loja", "5522981184419", "BOT_DETECTADO", "3492", "29/04/2026 09:21", 12),
    ("SH Tech", "5542988684665", "DESCARTADO", "2294", "29/04/2026 12:20", 12),
    ("Loja", "5528999261383", "DISPARO_REALIZADO", "3715", "04/05/2026 12:53", 7),
    ("Loja", "5527996616071", "BOT_DETECTADO", "3690", "05/05/2026 10:31", 6),
    ("TeamVL Celulares e Eletrônicos", "5551981740917", "FORMULARIO_ENVIADO", "2063", "05/05/2026 10:56", 6),
    ("Loja", "5531972481551", "BOT_DETECTADO", "4515", "05/05/2026 10:58", 6),
    ("Loja", "5531983790362", "DISPARO_REALIZADO", "4651", "05/05/2026 11:51", 6),
    ("Loja", "5533991632888", "BOT_DETECTADO", "4703", "06/05/2026 09:53", 5),
    ("Mix Case Sapucaia", "5551997911516", "BOT_DETECTADO", "2315", "11/05/2026 12:24", 0),
    ("Loja", "5554981285531", "BOT_DETECTADO", "2361", "11/05/2026 12:25", 0),
    ("Lojista", "5511995704410", "BOT_DETECTADO", "2505", "11/05/2026 12:26", 0),
    ("Loja", "5513988160091", "BOT_DETECTADO", "2606", "11/05/2026 12:26", 0),
    ("Loja", "5514997599851", "DISPARO_REALIZADO", "2655", "11/05/2026 12:27", 0),
    ("LéoCelulares", "5531992614263", "BOT_DETECTADO", "4592", "11/05/2026 12:30", 0),
    ("Loja", "5531993492330", "BOT_DETECTADO", "4601", "11/05/2026 12:30", 0),
    ("Loja", "5532998334968", "BOT_DETECTADO", "4677", "11/05/2026 12:31", 0),
    ("Loja", "5537999661992", "BOT_DETECTADO", "4883", "11/05/2026 12:32", 0),
    ("TOURO CELULARES", "5538997504466", "BOT_DETECTADO", "4927", "11/05/2026 12:33", 0),
    ("MeuCelular.com", "5512982343532", "BOT_DETECTADO", "2566", "11/05/2026 12:37", 0),
    ("Loja", "5554999270857", "NAO_QUALIFICADO", "4962", "11/05/2026 12:39", 0),
    ("Loja", "5555996555757", "BOT_DETECTADO", "4979", "11/05/2026 12:42", 0),
    ("Lojista", "5511988024256", "BOT_DETECTADO", "2455", "11/05/2026 12:46", 0),
    ("Loja", "5511956969600", "BOT_DETECTADO", "2297", "11/05/2026 13:18", 0),
]

# ─── ESTILOS ──────────────────────────────────────────────────────────────────

FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FILL_HEADER = PatternFill("solid", start_color="1F4E78")
FILL_CRITICAL = PatternFill("solid", start_color="FFE699")  # amarelo claro: 10+ dias parado
FILL_WARNING = PatternFill("solid", start_color="FFF2CC")   # creme: 5-9 dias
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def estilizar_header(sheet, num_cols):
    for col_idx in range(1, num_cols + 1):
        c = sheet.cell(row=1, column=col_idx)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    sheet.row_dimensions[1].height = 28
    sheet.freeze_panes = "A2"

def aplicar_zebra_e_destaque(sheet, num_cols, dias_col_idx):
    """Aplica linhas zebradas + destaque por dias parado."""
    for row_idx in range(2, sheet.max_row + 1):
        dias = sheet.cell(row=row_idx, column=dias_col_idx).value
        fill = None
        if isinstance(dias, int):
            if dias >= 10:
                fill = FILL_CRITICAL
            elif dias >= 5:
                fill = FILL_WARNING
        for col_idx in range(1, num_cols + 1):
            cel = sheet.cell(row=row_idx, column=col_idx)
            cel.font = Font(name="Arial", size=10)
            cel.alignment = ALIGN_LEFT
            cel.border = THIN_BORDER
            if fill:
                cel.fill = fill

# ─── PLANILHA ─────────────────────────────────────────────────────────────────

wb = Workbook()

# ===== Aba A — CADASTRO_COMPLETO =====
ws_a = wb.active
ws_a.title = "A_Cadastro_Completo"
ws_a.append(["#", "Loja / Nome", "Telefone", "Opp CRM", "Último contato", "Dias parado"])
for i, row in enumerate(LISTA_A, start=1):
    ws_a.append([i, *row])
estilizar_header(ws_a, 6)
aplicar_zebra_e_destaque(ws_a, 6, 6)
for col, width in zip("ABCDEF", [5, 38, 16, 10, 18, 14]):
    ws_a.column_dimensions[col].width = width
# Linha de instrução no topo (acima da tabela seria ideal, mas openpyxl força inserir)
ws_a.insert_rows(1)
ws_a["A1"] = "✅ AÇÃO: mover esses leads no CRM Evo Talks pra etapa 'Em Análise CAF' (stage 50). Trigger envia HSM link CAF + reforço biometria automaticamente."
ws_a["A1"].font = Font(name="Arial", size=11, bold=True, color="1F4E78")
ws_a["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_a.merge_cells("A1:F1")
ws_a.row_dimensions[1].height = 32

# ===== Aba B — AGUARDANDO_APROVACAO =====
ws_b = wb.create_sheet("B_Aguardando_Aprovacao")
ws_b.append(["#", "Loja / Nome", "Telefone", "Opp CRM", "Último contato", "Dias parado"])
for i, row in enumerate(LISTA_B, start=1):
    ws_b.append([i, *row])
estilizar_header(ws_b, 6)
aplicar_zebra_e_destaque(ws_b, 6, 6)
for col, width in zip("ABCDEF", [5, 38, 16, 10, 18, 14]):
    ws_b.column_dimensions[col].width = width
ws_b.insert_rows(1)
ws_b["A1"] = "🟠 AÇÃO: Eduardo/Nei revisar pré-análise e mover pra etapa 'Cadastro Recebido' (stage 49). Trigger envia HSM Complete o Cadastro pra VictorIA seguir com Fase 3."
ws_b["A1"].font = Font(name="Arial", size=11, bold=True, color="C65911")
ws_b["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_b.merge_cells("A1:F1")
ws_b.row_dimensions[1].height = 32

# ===== Aba C — ATENDIMENTO_HUMANO =====
ws_c = wb.create_sheet("C_Atendimento_Humano")
ws_c.append(["#", "Loja / Nome", "Telefone", "Status", "Opp CRM", "Último contato", "Dias parado"])
for i, row in enumerate(LISTA_C, start=1):
    ws_c.append([i, *row])
estilizar_header(ws_c, 7)
aplicar_zebra_e_destaque(ws_c, 7, 7)
for col, width in zip("ABCDEFG", [5, 38, 16, 22, 10, 18, 14]):
    ws_c.column_dimensions[col].width = width
ws_c.insert_rows(1)
ws_c["A1"] = "🚨 AÇÃO: lista priorizada por fase (avançados primeiro). Foco do Nei pra atender manualmente. 117 leads marcados com 'acionar_humano = true'."
ws_c["A1"].font = Font(name="Arial", size=11, bold=True, color="C00000")
ws_c["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_c.merge_cells("A1:G1")
ws_c.row_dimensions[1].height = 32

# ─── SAVE ─────────────────────────────────────────────────────────────────────
output_path = r"C:\Users\rocha\OneDrive\Área de Trabalho\TXT da Claude\acoes_aiva_2026-05-11.xlsx"
wb.save(output_path)
print(f"OK — arquivo salvo: {output_path}")
print(f"Lista A: {len(LISTA_A)} leads")
print(f"Lista B: {len(LISTA_B)} leads")
print(f"Lista C: {len(LISTA_C)} leads")
